import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

src = "/root/.claude/uploads/34498236-b66d-5586-88f6-a946b1dcc30e/f04950e1-HAVAIANAS__SC.md"
raw = open(src, encoding="utf-8").read()

# Remover escapes de markdown (\_ \[ \] \* \# etc.) que nao sao escapes validos de JSON
raw = re.sub(r'\\([_\[\]*#~`>])', r'\1', raw)

obj = json.loads(raw)
rows = obj["data"]
print("Linhas:", len(rows))

# Ordem das colunas: comecar pelas mais relevantes
preferred = ["marca", "bandeira", "tickets_amostra", "percentual_dimensao",
             "percentual_marca_dimensao", "oportunidade_dimensao"]
all_keys = list(rows[0].keys())
cols = preferred + [k for k in all_keys if k not in preferred]

wb = Workbook()
ws = wb.active
ws.title = "HAVAIANAS SC"

# Cabecalho
ws.append(cols)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")
for c in ws[1]:
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Dados
for r in rows:
    ws.append([r.get(k) for k in cols])

ws.freeze_panes = "A2"

# Largura de colunas
for i, k in enumerate(cols, 1):
    letter = ws.cell(row=1, column=i).column_letter
    maxlen = max([len(str(k))] + [len(str(r.get(k, ""))) for r in rows])
    ws.column_dimensions[letter].width = min(max(maxlen + 2, 10), 40)

out = "/home/user/hypr-claro/HAVAIANAS_SC.xlsx"
wb.save(out)
print("Salvo em:", out)
print("Colunas:", cols)
